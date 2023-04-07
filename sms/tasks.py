import csv
import json
import random
from datetime import timedelta
import requests
from celery import Celery
from celery.schedules import crontab
from celery.task import task, periodic_task
from celery.utils.log import get_task_logger
from django.db.models import Q
from openpyxl import load_workbook

from roberms.settings import BASE_DIR
from roberms_admin.models import Company, Appointment, EmalifyToken
from sms.models import *
from sms.utils import SDP, calculate_message_cost, get_access_token
import logging
logger = get_task_logger(__name__)
# localStorage = localStoragePy(BASE_DIR)

logging.basicConfig(filename="test.log", level=logging.DEBUG)


# @periodic_task(
#     run_every=(timedelta(seconds=3000)),
#     name="get_access_token",
#     ignore_result=True,
# )
# def get_access():
#     client_id = "dHq5I3HZDyR5vuh2wNhjGjSIBVCdNeie"
#     client_secret = "91B9moegcTJSCmtlOw4YmohnxXpJ6doZdxryVY12"
#     url = "https://api.emalify.com/v1/oauth/token"
#     headers = {
#         'Accept': 'application/json',
#         'Content-Type': 'application/json'
#     }
#     data = {
#         "client_id": client_id,
#         "client_secret": client_secret,
#         "grant_type": "client_credentials"
#     }
#     response = requests.post(url, json=data, headers=headers)
#     logging.error(response.text)
#     if EmalifyToken.objects.count() > 0:
#         mytoken = EmalifyToken.objects.first()
#         mytoken.token = json.loads(response.text)['access_token']
#         mytoken.save()
#     else:
#         EmalifyToken.objects.create(
#             token=json.loads(response.text)['access_token']
#         )
#     logger.info("Access token retrieved")


@task
def simple_sms_store(customer_id, total_message_cost, data, track_code):
    logging.error("Sms store got here")
    customer = Customer.objects.filter(id=customer_id).first()
    new_credit = customer.credit - total_message_cost
    customer.credit = new_credit
    customer.save()

    messages = []
    for a, b in data.items():
        s = ''.join(a.split())
        p = f"{254}{s[-9:]}"

        messages.append(OutgoingNew(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=p,
            text_message=b,
            track_code=track_code,
            sent_time=timezone.now(),
            usage_status=True,
            sender_type=customer.sender_type
        ))

        if len(messages) >= 500:
            OutgoingNew.objects.bulk_create(messages)
            messages.clear()
    if len(messages) > 0:
        OutgoingNew.objects.bulk_create(messages)
        messages.clear()
    return 'insertion complete'


@task()
def from_group_send(customer_id, total_message_cost, to_be_sent, trackingcode):
    logging.error("Sms store got here")
    customer = Customer.objects.filter(id=customer_id).first()
    new_credit = customer.credit - total_message_cost
    customer.credit = new_credit
    customer.save()

    messages = []
    for a, b in to_be_sent.items():
        # outgoing_new, created = OutgoingNew.objects.update_or_create(
        #     customer=customer,
        #     service_id=customer.service_id,
        #     access_code=customer.access_code,
        #     phone_number=a,
        #     text_message=b,
        #     track_code=trackingcode,
        #     sent_time=timezone.now(),
        #     usage_status=True,
        #     sender_type=customer.sender_type
        # )

        messages.append(
            OutgoingNew(
                customer=customer,
                service_id=customer.service_id,
                access_code=customer.access_code,
                phone_number=a,
                text_message=b,
                track_code=trackingcode,
                sent_time=timezone.now(),
                usage_status=True,
                sender_type=customer.sender_type
            )
        )
        if len(messages) > 500:
            OutgoingNew.objects.bulk_create(messages)
            messages.clear()
    if len(messages) > 0:
        OutgoingNew.objects.bulk_create(messages)
        messages.clear()
    return 'completed insertion'


@task()
def store_contact_task(group_id, extension, uploaded_file_url, f_path):
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        with open(file_path, 'r') as f:
            firstline = True
            for row in csv.reader(f):
                if firstline:
                    firstline = False
                    continue
                else:
                    # print(row[2])
                    p = f"{254}{row[1].replace(' ', '')[-9:]}"
                    Contact.objects.update_or_create(
                        name=row[0],
                        group_id=group_id,
                        phone_number=int(p),
                        email=row[2]
                    )
            CustomerTask.objects.filter(task_id=store_contact_task.id).update(
                status_complete=True
            )
            return 'completed'
    else:
        # print('work')
        # print(file_path)
        workbook = load_workbook(filename=f_path, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        for i in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=i, column=2).value != '':
                group_id = group_id
                name = worksheet.cell(row=i, column=1).value
                phone_number = str(worksheet.cell(row=i, column=2).value)
                email = worksheet.cell(row=i, column=3).value
                p = f"{254}{phone_number.replace(' ', '')[-9:]}"
                try:
                    Contact.objects.update_or_create(
                        name=name,
                        group_id=group_id,
                        phone_number=int(p),
                        email=email
                    )
                except TypeError:
                    continue
                except ValueError:
                    continue
                print('saved')
        # CustomerTask.objects.filter(task_id=store_contact_task.id).update(
        #     status_complete=True
        # )
        return 'completed'



@task()
def new_store_contact_task(group_id, extension, uploaded_file_url, f_path):
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        with open(file_path, 'r') as f:
            firstline = True
            contacts = []
            for row in csv.reader(f):
                if firstline:
                    firstline = False
                    continue
                else:
                    # print(row[2])
                    p = f"{254}{row[1].replace(' ', '')[-9:]}"
                    contact = Contact(
                        name=str(row[0]),
                        group_id=group_id,
                        phone_number=int(p),
                        email=row[2]
                    )
                    contacts.append(contact)
                    if len(contacts) >= 20000:
                        Contact.objects.bulk_create(contacts)
                        contacts.clear()
            Contact.objects.bulk_create(contacts)
            return 'completed'
    else:
        # print('work')
        # print(file_path)
        workbook = load_workbook(filename=f_path, read_only=False)
        worksheet = workbook[workbook.sheetnames[0]]
        contacts = []
        for i in range(2, worksheet.max_row):
            # if worksheet.cell(row=i, column=2).value != '':
            group_id = group_id
            name = worksheet.cell(row=i, column=1).value
            phone_number = str(worksheet.cell(row=i, column=2).value)
            email = worksheet.cell(row=i, column=3).value
            p = f"{254}{phone_number.replace(' ', '')[-9:]}"

            # Contact.objects.update_or_create(
            #     name=name,
            #     group_id=group_id,
            #     phone_number=int(p),
            #     email=email
            # )
            contact = Contact(
                name=name,
                group_id=group_id,
                phone_number=int(p),
                email=email
            )
            contacts.append(contact)
            if len(contacts) >= 20000:
                Contact.objects.bulk_create(contacts)
                contacts.clear()
            else:
                continue
        Contact.objects.bulk_create(contacts)
        return 'completed'


@periodic_task(
    run_every=(timedelta(seconds=60)),
    name="send_to_airtel",
    ignore_result=True,
)
def send_sms():
    access_token = get_access_token()
    url = "https://api.emalify.com/v1/projects/zyl6jnmp8yrvb93x/sms/simple/send"
    headers = {"Authorization": "Bearer %s" % access_token,
               'Accept': 'application/json',
               'Content-Type': 'application/json'
               }
    messages = OutgoingDone.objects.filter(Q(customer_id=105) | Q(customer_id=107),
                                           Q(delivery_status='UserNotExist') | Q(delivery_status='DeliveryImpossible'), extra_status=0)
    unique_code = random.randint(9000, 1000000)
    for message in messages:
        body = {
            "to": [message.phone_number],
            "message": message.text_message,
            "messageId": unique_code,
            "callback": "http://roberms.com",
            "from": "BIDCO_AFRCA",
        }
        response = requests.post(url, json=body, headers=headers)
        # message.request_identifier = json.loads(response.text)["data"][0]["messageId"]
        message.request_identifier = response.text
        message.extra_status = 1
        message.delivery_status = None
        message.save()
        logger.info(json.loads(response.text))
        # if json.loads(response.text)["status"] == "success":
    logger.info("send_to_airtel")


@periodic_task(
    run_every=(timedelta(seconds=60)),
    name="send_tibanet_sms",
    ignore_result=True,
)
def send_tibanet_sms():
    access_token = get_access_token()
    url = "https://api.emalify.com/v1/projects/zyl6jnmp8yrvb93x/sms/simple/send"
    headers = {"Authorization": "Bearer %s" % access_token,
               'Accept': 'application/json',
               'Content-Type': 'application/json'
               }
    messages = OutgoingDone.objects.filter(Q(customer_id=230),
                                           Q(delivery_status='UserNotExist') | Q(delivery_status='DeliveryImpossible'), extra_status=0)
    unique_code = random.randint(9000, 1000000)
    for message in messages:
        try:
            body = {
                "to": [message.phone_number],
                "message": message.text_message,
                "messageId": unique_code,
                "callback": "http://roberms.com",
                "from": "TIBANET",
            }
            response = requests.post(url, json=body, headers=headers)
            # message.request_identifier = json.loads(response.text)["data"][0]["messageId"]
            message.request_identifier = response.text
            message.extra_status = 1
            message.delivery_status = None
            message.save()
            logger.info(json.loads(response.text))
            # if json.loads(response.text)["status"] == "success":
        except Exception as e:
            logging.error(e)
    logger.info("send_to_airtel")


@periodic_task(
    run_every=(timedelta(seconds=60)),
    name="process_inbox",
    ignore_result=True,
)
def process_inbox():
    inbox = Inbox.objects.filter(processed=False)
    for message in inbox:
        message_list = message.message.split(sep=" ")
        hash_tag = message_list[0]
        tag = Tag.objects.filter(hashtag__icontains=hash_tag).first()
        if tag is not None:
            message.customer_id = tag.customer_id
            message.processed = True
            message.tag_id = tag.id
            message.message=message.message.replace(hash_tag, " ")
            message.created_at = timezone.now()
            message.updated_at = timezone.now()
            message.save()
            customer = tag.customer
            message_cost = calculate_message_cost(tag.response)
            customer_code = f"{datetime.datetime.today().date()}{customer.id}"

            if tag.response != "" and tag.response != " " and tag.response:
                if customer.credit >= message_cost:
                    result = OutgoingNew.objects.create(
                        phone_number=message.phone_number,
                        text_message=tag.response,
                        access_code=customer.sender_name,
                        customer_id=customer.id,
                        track_code=customer_code.replace('-', ''),
                        request_identifier=customer_code)
                    remaining_credit = customer.credit - message_cost
                    customer.credit = remaining_credit
                    customer.save()
        else:
            continue

    logger.info("Message processing complete")


@periodic_task(
    run_every=(timedelta(minutes=1)),
    name="get_delivery_status",
   ignore_result=True,
)
def get_delivery_status():
    access_token = get_access_token()
    print(access_token)
    messages = OutgoingDone.objects.filter(Q(customer_id=105)|Q(customer_id=107)|Q(customer_id=230)|Q(customer_id=7),
                                           Q(delivery_status__isnull=True)|Q(delivery_status='QUEUED'),
                                           request_identifier__isnull=False, extra_status=1)
    for message in messages:
        try:
            message_id = json.loads(message.request_identifier)["data"][0]["messageId"]
            d_url = f"https://api.emalify.com/v1/projects/zyl6jnmp8yrvb93x/sms/delivery-report?messageId={message_id}"
            headers = {
                "Authorization": "Bearer %s" % access_token,
                'Accept': 'application/json',
            }
            response = requests.get(d_url, headers=headers)
            message.delivery_status = json.loads(response.text)["data"][0]["status"]
            message.save()
        except Exception as e:
            logging.error(e)
    logger.info("Done getting delivery status")


@periodic_task(
    run_every=crontab(minute=0, hour='6'),
    name="check_appointments",
    ignore_result=True,
)
def check_appointments():
    sales_people = SalesPerson.objects.all()
    month = datetime.datetime.month
    year = datetime.datetime.year
    day = datetime.datetime.day
    customer = Customer.object.get(id=1)
    for sales_person in sales_people:
        unique_code = random.randint(9000, 1000000)
        companies = Company.objects.filter(sales_person_id=sales_person.id)
        message = f"Dear, {sales_person.first_name} {sales_person.last_name} your appointments for today are as follows: "
        for company in companies:
            appointment = Appointment.objects.filter(company_id=company.id, status_visited=False,
                                                      date_visited__year=year,
                                                      date_visited__month=month, date_visited__day=day).first()
            message += f"\n Company -> {company.name}, Appointment at {appointment.date_visited}."
        OutgoingNew.objects.create(
            phone_number=f"{254}{sales_person.phone_number.replace(' ', '')[-9:]}",
            text_message=message,
            access_code="ROBERMS_LTD",
            customer_id=1,
            track_code=unique_code,
            request_identifier=unique_code
        )
        message_cost = calculate_message_cost(message)
        remaining_credit = customer.credit - message_cost
        customer.credit = remaining_credit
        customer.save()
        message = ""

    logger.info("Done checking appointments")


@periodic_task(
    run_every=crontab(minute=0, hour='23'),
    name="check_for_new_sender_names",
    ignore_result=True,
)
def check_for_new_sender_names():
    url = "https://sms.procom.co.ke/sms/v1/procom/company/sender/names"
    for customer in Customer.objects.all():
        body = {
            'sender_name': customer.access_code,
            'company_id': 1
        }
        response = requests.post(url=url, data=body)
        logging.info(f"{response.text} {response.status_code}")


@periodic_task(
    run_every=(timedelta(seconds=60)),
    name="send_airtel_sms",
    ignore_result=True,
)
def send_airtel_sms():
    access_token = get_access_token()
    url = "https://api.emalify.com/v1/projects/zyl6jnmp8yrvb93x/sms/simple/send"
    headers = {"Authorization": "Bearer %s" % access_token,
               'Accept': 'application/json',
               'Content-Type': 'application/json'
               }

    messages = OutgoingDone.objects.filter(Q(access_code='CLEANFAX')|Q(access_code='MysafeVault'),
                                           Q(delivery_status='UserNotExist') | Q(delivery_status='DeliveryImpossible'), extra_status=0,
                                           sent_time__gt="2021-02-24 23:59:59.999999")
    unique_code = random.randint(9000, 1000000)
    for message in messages:
        body = {
            "to": [message.phone_number],
            "message": message.text_message,
            "messageId": unique_code,
            "callback": "http://roberms.com",
            "from": "EMALIFY",
        }
        response = requests.post(url, json=body, headers=headers)
        # message.request_identifier = json.loads(response.text)["data"][0]["messageId"]
        message.request_identifier = response.text
        message.extra_status = 1
        message.delivery_status = None
        message.save()
        logger.info(json.loads(response.text))
        # if json.loads(response.text)["status"] == "success":
    logger.info("send_to_airtel")


@periodic_task(
    run_every=(timedelta(seconds=60)),
    name="send_airtel_apex",
    ignore_result=True,
)
def send_airtel_apex():
    access_token = get_access_token()
    url = "https://api.emalify.com/v1/projects/zyl6jnmp8yrvb93x/sms/simple/send"
    headers = {"Authorization": "Bearer %s" % access_token,
               'Accept': 'application/json',
               'Content-Type': 'application/json'
               }

    messages = OutgoingDone.objects.filter(Q(delivery_status='UserNotExist') | Q(delivery_status='DeliveryImpossible'),
                                           extra_status=0, access_code='APEX_STEEL',
                                           sent_time__gt="2021-02-24 23:59:59.999999")
    unique_code = random.randint(9000, 1000000)
    for message in messages:
        body = {
            "to": [message.phone_number],
            "message": message.text_message,
            "messageId": unique_code,
            "callback": "http://roberms.co.ke",
            "from": "APEX_STEEL",
        }
        response = requests.post(url, json=body, headers=headers)
        # message.request_identifier = json.loads(response.text)["data"][0]["messageId"]
        message.request_identifier = response.text
        message.extra_status = 1
        message.delivery_status = None
        message.save()
        logger.info(json.loads(response.text))
        # if json.loads(response.text)["status"] == "success":
    logger.info("apex_send_to_airtel")


@periodic_task(
    run_every=(timedelta(seconds=60)),
    name="send_airtel_tom_water",
    ignore_result=True,
)
def send_airtel_tom_water():
    access_token = get_access_token()
    url = "https://api.emalify.com/v1/projects/zyl6jnmp8yrvb93x/sms/simple/send"
    headers = {"Authorization": "Bearer %s" % access_token,
               'Accept': 'application/json',
               'Content-Type': 'application/json'
               }

    messages = OutgoingDone.objects.filter(Q(delivery_status='UserNotExist') | Q(delivery_status='DeliveryImpossible'),
                                           extra_status=0, access_code='Tom_Water',
                                           sent_time__gt="2021-04-11 23:59:59.999999")
    unique_code = random.randint(9000, 1000000)
    for message in messages:
        body = {
            "to": [message.phone_number],
            "message": message.text_message,
            "messageId": unique_code,
            "callback": "http://roberms.co.ke",
            "from": "Tom_water",
        }
        response = requests.post(url, json=body, headers=headers)
        # message.request_identifier = json.loads(response.text)["data"][0]["messageId"]
        message.request_identifier = response.text
        message.extra_status = 1
        message.delivery_status = None
        message.save()
        logger.info(json.loads(response.text))
        # if json.loads(response.text)["status"] == "success":
    logger.info("apex_send_to_tom_water")


@task()
def clean_group_contacts(group_id, client_id):
    client = Customer.objects.get(id=client_id)
    messages = OutgoingDone.objects.raw(
        f"SELECT DISTINCT  `id`, `phone_number` FROM `sms_outgoingdone` WHERE `access_code` = '{client.access_code}' "
        f"and `customer_id` = '{client.id}' "
        f"and (`delivery_status` = 'SenderName Blacklisted' or `delivery_status` = 'AbsentSubscriber' or `delivery_status` = 'DeliveryImpossible')"
    )
    phone_numbers = []

    try:
        for message in messages:
            phone_numbers.append(message.phone_number)

            if len(list(set(phone_numbers))) >= 1000:
                contacts = Contact.objects.filter(group_id=group_id, phone_number__in=list(set(phone_numbers)))
                contacts.delete()
                phone_numbers.clear()

        contacts = Contact.objects.filter(group_id=group_id, phone_number__in=list(set(phone_numbers)))
        contacts.delete()
        phone_numbers.clear()

        logging.info('Contacts Cleaned')
    except Exception as e:
        logging.error(f'Contacts Cleaning Error {e}')