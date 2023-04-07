from django.shortcuts import render, redirect
from django.contrib import messages
from celery.task import task
from django.utils import timezone
from openpyxl import load_workbook

from sms.views import login_required, is_user_customer, Merged
from sms.models import Customer, Group, CustomerSubAccounts, Contact, OutgoingNew
from sms.utils import get_message_parameters, get_excel_content, get_parameter_column, get_phone_number_column, \
    BulkCreateManager
import random
from sms.utils import calculate_message_cost
from .forms import SchedulePerMonthForm
from .tasks import store_scheduled_sms
from .models import ScheduledMessage, SchedulePerMonth


@login_required()
@is_user_customer
def schedule_sms(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer:
        pass
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner

    if customer is not None:
        groups = Group.objects.filter(customer=customer)
        if request.method == 'POST':
            message = request.POST.get('Message')
            context = {
                'message': message,
                'group': request.POST['group'],
                'scheduled_time': request.POST['scheduled_time']
            }
            request.session['c_data'] = context
            return redirect('scheduled_sms:sample_merged')
        else:
            context = {
                'groups': groups
            }
            return render(request, 'scheduled_sms/schedule_sms.html', context)
    else:
        messages.error(request, 'An error occured. Please Try Again!!')
        # context = {
        #     'groups': groups
        # }
        return render(request, 'scheduled_sms/schedule_sms.html')


@login_required()
@is_user_customer
def scheduled_sample_merged(request):
    data = request.session.get('c_data')
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer:
        pass
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner

    if customer is not None:
        parameters = get_message_parameters(message=data['message'])
        if 'group' in data:
            c_group = Group.objects.filter(id=data['group']).first()
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for contact in Contact.objects.filter(group_id=request.POST['group'], is_active=True):
                    complete_message = message
                    if parameters:
                        for parameter in parameters:
                            if parameter == 'Name':
                                complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                    else:
                        to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/schedule_sms.html')
                else:
                    # ScheduledMessage.objects.create(
                    #     customer_id=customer.id,
                    #     text_message=message,
                    #     sender_name=customer.access_code,
                    #     track_code=trackingcode,
                    #     scheduled_time=data['scheduled_time'],
                    #     group_id=c_group.id
                    # )
                    store_scheduled_sms.delay(customer.id, total_message_cost, to_be_sent, trackingcode, data['scheduled_time'])
                    messages.success(request, 'Message scheduled successfully')
                    return redirect('scheduled_sms:list_scheduled_sms')
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'group':
                        contacts = Contact.objects.filter(group_id=b)
                        for contact in contacts:
                            complete_message = data['message']
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': c_group.id
                }
                # pprint(context)
                return render(request, 'scheduled_sms/schedule_merged.html', context)
    else:
        messages.error(
            request,
            'You do not have enough credit to make this request, kindly recharge to proceed')
        return render(request, 'scheduled_sms/schedule_sms.html')


@login_required()
@is_user_customer
def list_scheduled_sms(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    text_messages = ScheduledMessage.objects.filter(customer_id=customer.id).order_by('-created_at')

    context = {
        "text_messages": text_messages
    }
    return render(request, 'scheduled_sms/list.html', context)


@login_required()
@is_user_customer
def list_scheduled_per_month(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer:
        pass
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner

    text_messages = SchedulePerMonth.objects.filter(customer_id=customer.id)
    context = {
        "text_messages": text_messages
    }
    return render(request, 'schedule_per_month/list.html', context)


@login_required()
@is_user_customer
def schedule_per_month(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer:
        pass
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner

    if request.method == 'POST':
        form = SchedulePerMonthForm(request.POST)

        if form.is_valid():
            text_message = form.save(commit=False)
            text_message.is_active = True
            text_message.phone_number = f"{254}{request.POST['phone_number'].replace(' ', '')[-9:]}"
            text_message.save()
            message = f'Message Scheduled Successfully For {text_message.phone_number} To Be Sent Monthly ' \
                      f'On the {text_message.send_on} Of Each Month'
            messages.success(request, message)
            return redirect("scheduled_sms:list_scheduled_per_month")
        else:
            context = {
                "customer": customer,
                "form": form
            }
            print(form)
            messages.error(request, 'Invalid form data')
            return render(request, 'schedule_per_month/create.html', context)

    context = {
        "customer": customer
    }
    return render(request, 'schedule_per_month/create.html', context)


@login_required()
@is_user_customer
def edit_schedule_per_month(request, id):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer:
        pass
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner

    scheduled_message = SchedulePerMonth.objects.filter(id=id, customer_id=customer.id).first()
    if request.method == 'POST':
        form = SchedulePerMonthForm(request.POST, instance=scheduled_message)

        if form.is_valid():
            text_message = form.save(commit=False)
            text_message.phone_number = f"{254}{request.POST['phone_number'].replace(' ', '')[-9:]}"
            text_message.save()
            message = f'Message Scheduled Successfully For {form.phone_number} To Be Sent Monthly ' \
                      f'On the {form.send_on} Of Each Month'
            messages.success(request, message)
            return redirect("scheduled_sms:list_scheduled_per_month")
        else:
            context = {
                "customer": customer,
                "scheduled_message": scheduled_message,
                "form": form
            }
            messages.error(request, 'Invalid form data')
            return render(request, 'schedule_per_month/edit.html', context)

    context = {
        "customer": customer,
        "scheduled_message": scheduled_message
    }
    return render(request, 'schedule_per_month/edit.html', context)


@login_required()
@is_user_customer
def deactivate_scheduled_per_month(request, id):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer:
        pass
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner

    scheduled = SchedulePerMonth.objects.filter(id=id, customer_id=customer.id).first()
    if scheduled:
        if scheduled.is_active:
            scheduled.is_active = False
            scheduled.save()
            messages.success(request, 'Scheduled Message De-activated')
        elif not scheduled.is_active:
            scheduled.is_active = True
            scheduled.save()
            messages.success(request, 'Scheduled Message Activated')
        return redirect("scheduled_sms:list_scheduled_per_month")
    return redirect("scheduled_sms:list_scheduled_per_month")


@login_required()
@is_user_customer
def schedule_sms_dashboard(request):
    return render(request, 'scheduled_sms/dashboard.html')


@login_required()
@is_user_customer
def simple_sms(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        if request.method == 'POST':
            message = request.POST.get('text_message')
            scheduled_time = request.POST.get('scheduled_time')
            phone_numbers = request.POST.get('phone_numbers').split(',')
            data = {}
            print(phone_numbers)
            new_phone_numbers = []
            for phone_number in phone_numbers:
                phone_number = f"{254}{phone_number[-9:]}"
                if phone_number != '':
                    new_phone_numbers.append(phone_number)
            message_cost = calculate_message_cost(message)
            total_message_cost = message_cost * len(new_phone_numbers)

            if customer.credit >= total_message_cost:
                for p in new_phone_numbers:
                    data[p] = message

                request.session['simple_messages'] = data
                request.session['scheduled_time'] = scheduled_time
                return redirect("scheduled_sms:simple_sms_preview")
            else:
                messages.error(request, 'You do not have enough credit to perform this action. Kindly Top Up To Continue')
                return redirect("scheduled_sms:simple_sms")
        else:
            context = {
                'customer': customer
            }
            return render(request, 'scheduled_sms/simple/schedule_sms.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        if request.method == 'POST':
            message = request.POST.get('text_message')
            phone_numbers = request.POST.get('phone_numbers').splitlines()
            scheduled_time = request.POST.get('scheduled_time')

            data = {}
            new_phone_numbers = []
            for phone_number in phone_numbers:
                phone_number = f"{254}{phone_number[-9:]}"
                if phone_number != '':
                    new_phone_numbers.append(phone_number)
            message_cost = calculate_message_cost(message)
            total_message_cost = message_cost * len(new_phone_numbers)

            if customer.credit >= total_message_cost:
                for p in new_phone_numbers:
                    data[p] = message
                request.session['simple_messages'] = data
                request.session['scheduled_time'] = scheduled_time
                return redirect("scheduled_sms:simple_sms_preview")
            else:
                messages.error(request,
                               'You do not have enough credit to perform this action. Kindly Top Up To Continue')
                return redirect("scheduled_sms:simple_sms")
        else:
            context = {
                'customer': customer
            }
            return render(request, 'scheduled_sms/simple/schedule_sms.html', context)


@login_required()
@is_user_customer
def simple_sms_preview(request):
    context = {
        'data': request.session['simple_messages']
    }
    return render(request, 'scheduled_sms/simple/sample_merged.html', context)


@login_required()
@is_user_customer
def send(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        if request.method == 'POST':
            data = request.session['simple_messages']
            scheduled_time = request.session['scheduled_time']

            total_message_cost = 0
            for a, b in data.items():
                message_cost = calculate_message_cost(b)
                total_message_cost += message_cost

            if customer.credit >= total_message_cost:
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)
                track_code = trackingcode

                store_scheduled_sms.delay(customer.id, total_message_cost, data, trackingcode, scheduled_time)
                # simple_sms_store.delay(customer.id, total_message_cost, data, track_code)

                request.session.delete("simple_messages")
                request.session.delete("scheduled_time")
                return redirect('scheduled_sms:list_scheduled_sms')
            else:
                messages.error(request, 'You do not have enough credit in your account to make this request. '
                                        'Please Recharge To Continue')
                context = {
                    'phone_numbers': request.POST.get('phone_numbers'),
                    'message': request.POST.get('text_message'),
                    'customer': customer
                }
                return render(request, 'scheduled_sms/simple/schedule_sms.html', context)
        else:
            context = {
                'customer': customer
            }
            return render(request, 'scheduled_sms/simple/schedule_sms.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        credit = customer.credit
        if request.method == 'POST':
            data = request.session['simple_messages']
            scheduled_time = request.session['scheduled_time']

            total_message_cost = 0
            for a, b in data.items():
                message_cost = calculate_message_cost(b)
                total_message_cost += message_cost

            if customer.credit >= total_message_cost:
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)
                track_code = trackingcode

                store_scheduled_sms.delay(customer.id, total_message_cost, data, trackingcode, scheduled_time)
                # simple_sms_store.delay(customer.id, total_message_cost, data, track_code)

                request.session.delete("simple_messages")
                request.session.delete("scheduled_time")
                return redirect('scheduled_sms:list_scheduled_sms')
            else:
                messages.error(request, 'You do not have enough credit in your account to make this request. '
                                        'Please Recharge To Continue')
                context = {
                    'phone_numbers': request.POST.get('phone_numbers'),
                    'message': request.POST.get('text_message'),
                    'customer': customer
                }
                return render(request, 'scheduled_sms/simple/schedule_sms.html', context)
        else:
            context = {
                'customer': customer
            }
            return render(request, 'scheduled_sms/simple/schedule_sms.html', context)


@login_required()
@is_user_customer
def schedule_excel(request):
    if request.method == 'POST' and request.FILES['myfile']:
        file = request.FILES['myfile']
        context = get_excel_content(file)
        return render(request, 'scheduled_sms/from_file/schedule.html', context)
    return render(request, 'scheduled_sms/from_file/schedule.html')


@login_required()
@is_user_customer
def schedule_excel_merged(request):
    data = []
    if request.method == 'POST':
        message = request.POST['Message']
        scheduled_date = request.POST['schedule_date']
        phone_number_field = request.POST['NumberField']
        file = request.POST['file_path']

        file_path = file.split('/', 1)[1]
        workbook = load_workbook('media/%s' % file_path)
        sheet_names = workbook.sheetnames

        sheet = sheet_names[0]
        worksheet = workbook.get_sheet_by_name(sheet)

        parameters = get_message_parameters(message=message)
        parameter_cells = get_parameter_column(parameters=parameters, worksheet=worksheet)
        phone_number_column = get_phone_number_column(phone_number_field=phone_number_field, worksheet=worksheet)

        max_row = worksheet.max_row
        max_column = worksheet.max_column
        for i in range(2, max_row + 1):
            person_message = {}
            new_message = message
            sms = ''
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)

                for a, b in parameter_cells.items():
                    if j == b:
                        new_message = new_message.replace('[%s]' % a, str(cell_obj.value))
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)
                if j == phone_number_column:
                    phone_number = cell_obj.value
                    sms = Merged(phone_number, new_message)
                    person_message['phone_number'] = phone_number
                    person_message['message'] = new_message
            data.append(sms)
        data_dict = {}
        for d in data:
            data_dict[d.phone_number] = d.message
        request.session['data'] = data_dict
        request.session['excel_scheduled_date'] = scheduled_date
        return render(request, 'scheduled_sms/from_file/sample_merged.html', {'data': data})


@login_required()
@is_user_customer
def schedule_excel_confirm(request):
    if request.method == 'POST':
        tracking_code = random.randint(1, 1000000)
        while OutgoingNew.objects.filter(track_code=tracking_code).count() > 0:
            tracking_code = random.randint(1, 1000000)

        c_messages = request.session['data']
        scheduled_date = request.session['excel_scheduled_date']
        customer = Customer.objects.filter(user_ptr_id=request.user).first()
        if customer is not None:
            actual_messages_cost = 0
            for a, message in c_messages.items():
                message_cost = calculate_message_cost(message)
                actual_messages_cost += message_cost

            if customer.credit >= actual_messages_cost:
                bulk_mgr = BulkCreateManager(chunk_size=100)
                for a, message in c_messages.items():
                    p = f"{254}{a.replace(' ', '')[-9:]}"

                    bulk_mgr.add(ScheduledMessage(
                        customer=customer,
                        sender_name=customer.access_code,
                        phone_number=p,
                        text_message=message,
                        track_code=tracking_code,
                        scheduled_time=scheduled_date,
                        usage_status=True,
                        sender_type=customer.sender_type
                    ))

                bulk_mgr.done()
                customer.credit = customer.credit - actual_messages_cost
                customer.save()
                request.session.delete("data")
                request.session.delete("excel_scheduled_date")
                return redirect('scheduled_sms:list_scheduled_per_track_code')
            else:
                data = []
                data_dict = request.session.get('data')
                for a, b in data_dict.items():
                    sms = Merged(a, b)
                    data.append(sms)
                messages.warning(request,
                                 'You do not have enough credit in your account to perform this action please recharge to continue')
                return render(request, 'scheduled_sms/from_file/sample_merged.html', {'data': data})
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user).first().owner
            actual_messages_cost = 0
            for a, message in c_messages.items():
                message_cost = calculate_message_cost(message)
                actual_messages_cost += message_cost

            if customer.credit >= actual_messages_cost:
                bulk_mgr = BulkCreateManager(chunk_size=100)
                for a, message in c_messages.items():
                    p = f"{254}{a.replace(' ', '')[-9:]}"

                    bulk_mgr.add(ScheduledMessage(
                        customer=customer,
                        sender_name=customer.access_code,
                        phone_number=p,
                        text_message=message,
                        track_code=tracking_code,
                        scheduled_time=scheduled_date,
                        usage_status=True,
                        sender_type=customer.sender_type
                    ))
                bulk_mgr.done()
                customer.credit = customer.credit - actual_messages_cost
                customer.save()
                request.session.delete("data")
                request.session.delete("excel_scheduled_date")
                return redirect('scheduled_sms:list_scheduled_per_track_code')
            else:
                data = []
                data_dict = request.session.get('data')
                for a, b in data_dict.items():
                    sms = Merged(a, b)
                    data.append(sms)
                messages.warning(request,
                                 'You do not have enough credit in your account to perform this action please recharge to continue')
                return render(request, 'scheduled_sms/from_file/sample_merged.html', {'data': data})
    return render(request, 'sms/result.html')


# @login_required()
# @is_user_customer
# def list_scheduled_messages(request):
#     customer = Customer.objects.filter(user_ptr_id=request.user).first()
#     messages = ScheduledMessage.objects.filter(customer_id=customer.id)
#     context = {
#         "text_messages": messages
#     }
#     return render(request, 'scheduled_sms/from_file/list.html', context)


@login_required()
@is_user_customer
def list_scheduled_per_track_code(request):
    customer = Customer.objects.filter(user_ptr_id=request.user).first()
    text_messages = ScheduledMessage.objects.values(
        'track_code', 'sender_name', 'scheduled_time', 'is_sent', 'active'
    ).filter(customer_id=customer.id).distinct()
    context = {
        "text_messages": text_messages
    }
    return render(request, 'scheduled_sms/list_per_track_code.html', context)


@login_required()
@is_user_customer
def detail_scheduled_per_track_code(request, track_code):
    customer = Customer.objects.filter(user_ptr_id=request.user).first()
    text_messages = ScheduledMessage.objects.values(
        'track_code', 'sender_name', 'scheduled_time', 'is_sent', 'active'
    ).filter(customer_id=customer.id, track_code=track_code)
    context = {
        "text_messages": text_messages,
        "track_code": track_code
    }
    return render(request, 'scheduled_sms/list_per_track_details.html', context)


@login_required()
@is_user_customer
def disable_scheduled_per_track_code(request, track_code):
    customer = Customer.objects.filter(user_ptr_id=request.user).first()
    text_messages = ScheduledMessage.objects.filter(customer_id=customer.id, track_code=track_code)
    first_text = text_messages.first()
    if first_text.active:
        text_messages.update(active=False)
        messages.success(request, f'Messages with unique reference: {track_code} disabled successfully')
        return redirect('scheduled_sms:list_scheduled_per_track_code')
    elif not first_text.active:
        text_messages.update(active=True)
        messages.success(request, f'Messages with unique reference: {track_code} activated successfully')
        return redirect('scheduled_sms:list_scheduled_per_track_code')

